"""FRPS (Federal Reserve Payments Study) ingestion.

See docs/source_contracts/frps.md for the full contract. Downloads the
published top-line data workbook and reshapes it from the Fed's wide,
indented "Table 1" layout (payment type x [2015, 2018, 2021, 2024] x
[Number, Value, Average]) into a long Bronze table at grain
(row_number, collection_year).

The workbook's row hierarchy (Total > Cards > Debit cards > Non-prepaid...)
is encoded via Excel cell indentation, not an explicit parent column, and
labels repeat at different levels (e.g. two "Credit transfers" rows under
different parents) -- so a flat `payment_type` string alone can't be a
unique or unambiguous key. `category_path` reconstructs the full breadcrumb
from indentation so downstream consumers can group at whatever level of the
hierarchy they need without guessing.

Usage:
    python frps_ingest.py [--url <workbook-url>]
"""
from __future__ import annotations

import argparse
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import requests

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from common.lineage import add_lineage, write_landed_parquet  # noqa: E402
from common.logging_setup import get_logger  # noqa: E402

logger = get_logger("frps_ingest")

DEFAULT_URL = "https://www.federalreserve.gov/publications/images/FRPS_CY2024_IDR_data.xlsx"
SHEET_NAME = "Table 1"
SOURCE_SYSTEM = "frps_ingest"
OUTPUT_TABLE = "frps_payment_volumes"

# (collection_year, number_col, value_col, avg_col) -- 1-indexed columns in
# "Table 1"; verified by inspecting the actual downloaded workbook, not
# guessed from the landing-page description.
YEAR_BLOCKS = [
    (2015, 3, 4, 5),
    (2018, 6, 7, 8),
    (2021, 9, 10, 11),
    (2024, 12, 13, 14),
]
LABEL_COL = 2
DATA_START_ROW = 6


def _download(url: str, dest: Path) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    dest.write_bytes(resp.content)
    return dest


def _parse_workbook(path: Path) -> pd.DataFrame:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME]

    rows = []
    path_stack: dict[int, str] = {}  # indent_level -> label at that level

    for r in range(DATA_START_ROW, ws.max_row + 1):
        label_cell = ws.cell(row=r, column=LABEL_COL)
        label = label_cell.value
        if label is None:
            continue

        # Rows with no numeric data at all are section headers (e.g.
        # "Additional estimates") or the trailing "Note:" text block -- skip.
        has_data = any(
            ws.cell(row=r, column=c).value is not None
            for _, n, v, a in YEAR_BLOCKS
            for c in (n, v, a)
        )
        if not has_data:
            continue

        indent = int((label_cell.alignment.indent or 0))
        level = indent // 2
        path_stack[level] = str(label).strip()
        # Drop deeper levels that no longer apply now that we've moved to a shallower one.
        for deeper in [lvl for lvl in path_stack if lvl > level]:
            del path_stack[deeper]
        category_path = " > ".join(path_stack[lvl] for lvl in sorted(path_stack))

        for collection_year, num_col, val_col, avg_col in YEAR_BLOCKS:
            number_billions = ws.cell(row=r, column=num_col).value
            value_trillions = ws.cell(row=r, column=val_col).value
            avg_amount = ws.cell(row=r, column=avg_col).value
            if number_billions is None and value_trillions is None:
                continue
            rows.append({
                "row_number": r,
                "hierarchy_level": level,
                "payment_type_label": str(label).strip(),
                "category_path": category_path,
                "collection_year": collection_year,
                "count_billions": number_billions,
                "value_trillions_usd": value_trillions,
                "avg_transaction_amount_usd": avg_amount,
            })

    return pd.DataFrame(rows)


def main(url: str = DEFAULT_URL) -> None:
    run_id = str(uuid.uuid4())
    tmp_path = Path(__file__).resolve().parent / "_downloads" / "frps_workbook.xlsx"

    logger.info("Downloading FRPS workbook from %s", url)
    _download(url, tmp_path)

    df = _parse_workbook(tmp_path)
    df["source_url"] = url
    df["publication_date"] = None  # not machine-readable from the workbook; see docs/source_contracts/frps.md
    df["pulled_at"] = datetime.now(timezone.utc).isoformat()

    landed = add_lineage(df, SOURCE_SYSTEM, run_id)
    out_path = write_landed_parquet(landed, str(PROJECT_ROOT / "data" / "bronze"), OUTPUT_TABLE)

    logger.info("Landed %d rows -> %s", len(landed), out_path)
    logger.info("Rows by collection_year:\n%s", df.groupby("collection_year").size().to_string())


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ingest the FRPS top-line payment volume workbook.")
    parser.add_argument("--url", default=DEFAULT_URL)
    args = parser.parse_args()
    main(args.url)
